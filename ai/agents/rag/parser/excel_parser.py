import io
import openpyxl
from .parser_interface import BaseParser

class ExcelParser(BaseParser):
    def parse(self, file_bytes: bytes, file_name: str) -> str:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            output = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                output.append(f"### Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    if not any(row):
                        continue
                    row_str = [str(cell) if cell is not None else "" for cell in row]
                    output.append(" | ".join(row_str))
                output.append("")
            return "\n".join(output)
        except Exception as e:
            return f"Failed to parse Excel file {file_name}: {str(e)}"
